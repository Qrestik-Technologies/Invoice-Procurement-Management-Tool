"""Excel export service.

Builds an .xlsx workbook from invoice query results and returns bytes.
Uses openpyxl (no external dependency beyond pip install openpyxl).
"""
import io
from datetime import datetime
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.domain import Invoice


_BRAND_BLUE = "FF0C447C"
_LIGHT_GREY = "FFF3F4F6"
_HEADERS = [
    "Invoice #", "Customer", "Status", "Amount", "Currency",
    "Issue Date", "Due Date", "Received At", "Description",
]


def export_invoices_to_excel(invoices: Sequence[Invoice]) -> bytes:
    """Return a .xlsx file as bytes for the given invoice list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Header row
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=_BRAND_BLUE)
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    alt_fill = PatternFill("solid", fgColor=_LIGHT_GREY)
    for row_idx, inv in enumerate(invoices, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        row_data = [
            inv.invoice_number,
            inv.customer.name if inv.customer else inv.customer_id,
            inv.status.value,
            float(inv.amount),
            inv.currency,
            inv.issue_date.isoformat() if inv.issue_date else None,
            inv.due_date.isoformat() if inv.due_date else None,
            inv.received_at.isoformat() if inv.received_at else None,
            inv.description,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill

    # Auto-width columns
    for col_idx in range(1, len(_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(invoices) + 2)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
